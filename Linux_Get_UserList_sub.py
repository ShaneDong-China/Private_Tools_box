# -*- coding: utf-8 -*-
"""
获取Linux用户列表工具 (PySide6)
"""
import sys, os, logging, threading
from datetime import datetime

import paramiko
import pandas as pd
from openpyxl import load_workbook

from PySide6.QtWidgets import (QWidget, QApplication, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QFileDialog, QMessageBox)
from PySide6.QtCore import QTimer
from PySide6.QtGui import QFont

logger = logging.getLogger(__name__)


class UserListFetcher(QWidget):
    def __init__(self):
        super().__init__()
        self._init_logging()
        self.setWindowTitle("获取Linux用户列表")
        self.resize(800, 600)
        self.setStyleSheet("""
            UserListFetcher { background:#f5f6fa; }
            QLineEdit { border:none; border-bottom:1px solid #dfe6e9; padding:6px 4px; background:transparent; font-size:13px; }
            QLineEdit:focus { border-bottom:2px solid #0984e3; }
            QTextEdit { border:1px solid #dfe6e9; border-radius:4px; background:white; font-size:13px; }
        """)
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(24, 20, 24, 20)

        title = QLabel("获取Linux用户列表")
        title.setStyleSheet("font-size:18px;font-weight:bold;color:#2d3436;")
        layout.addWidget(title)

        l1 = QHBoxLayout()
        l1.addWidget(QLabel("Excel 文件"))
        self.fe = QLineEdit()
        self.fe.setPlaceholderText("选择包含 host,port,username,password 的Excel文件")
        l1.addWidget(self.fe)
        btn = QPushButton("浏览")
        btn.setStyleSheet("QPushButton{background:#0984e3;color:white;padding:7px 20px;border:none;border-radius:4px;font-size:13px;}QPushButton:hover{background:#0873c4;}")
        btn.clicked.connect(lambda: self.fe.setText(
            QFileDialog.getOpenFileName(self, "选择Excel","","Excel (*.xlsx)")[0]))
        l1.addWidget(btn)
        layout.addLayout(l1)

        l2 = QHBoxLayout()
        l2.addWidget(QLabel("输出列号 (默认 J 列)"))
        self.col_input = QLineEdit("J")
        self.col_input.setFixedWidth(50)
        l2.addWidget(self.col_input)
        l2.addStretch()
        layout.addLayout(l2)

        btn_row = QHBoxLayout()
        self.run_btn = QPushButton("开始获取")
        self.run_btn.setStyleSheet("QPushButton{background:#27ae60;color:white;font-weight:bold;padding:10px 36px;border:none;border-radius:4px;font-size:14px;}QPushButton:hover{background:#219a52;}QPushButton:disabled{background:#b2bec3;}")
        self.run_btn.clicked.connect(self._start)
        btn_row.addWidget(self.run_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setFont(QFont("Consolas", 10))
        layout.addWidget(self.log)

    def _init_logging(self):
        log_dir = os.path.join(os.path.dirname(__file__), "logs")
        os.makedirs(log_dir, exist_ok=True)
        fh = logging.FileHandler(
            os.path.join(log_dir, f"获取用户列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"),
            encoding="utf-8")
        fh.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(fh)
        logger.setLevel(logging.INFO)

    def _log(self, msg):
        logger.info(msg)
        self.log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
        self.log.ensureCursorVisible()

    def _start(self):
        fp = self.fe.text()
        if not fp:
            QMessageBox.warning(self, "提示", "请选择Excel文件")
            return
        self.run_btn.setEnabled(False)
        self.run_btn.setText("获取中...")
        self.log.clear()
        self._log("开始获取用户列表...")
        col_letter = self.col_input.text().strip().upper() or "J"
        threading.Thread(target=self._run, args=(fp, col_letter), daemon=True).start()

    def _run(self, fp, col_letter):
        try:
            df = pd.read_excel(fp)
            col_idx = ord(col_letter) - ord('A') + 1
            ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            wb = load_workbook(fp)
            ws = wb.active

            for i, (_, row) in enumerate(df.iterrows()):
                host = str(row.iloc[0])
                port = int(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else 22
                user = str(row.iloc[2]) if len(row) > 2 else ""
                pwd = str(row.iloc[3]) if len(row) > 3 else ""
                self._log(f"连接 {host}...")
                try:
                    ssh = paramiko.SSHClient()
                    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    ssh.connect(host, port=port, username=user, password=pwd, timeout=10)
                    _, stdout, _ = ssh.exec_command(
                        """awk -F: '$3>=1000 && $7!="/sbin/nologin" && $7!="/usr/sbin/nologin" {print $1}' /etc/passwd""")
                    users = stdout.read().decode().splitlines()
                    ssh.close()

                    row_excel = i + 2
                    ws.cell(row=row_excel, column=col_idx, value=ts)
                    for j, u in enumerate(users[:20]):
                        ws.cell(row=row_excel, column=col_idx + 1 + j, value=u)
                    self._log(f"✅ {host} 获取到 {len(users)} 个用户")
                except Exception as e:
                    self._log(f"❌ {host} 失败: {e}")

            wb.save(fp)
            self._log(f"完成！结果已写入 {fp}")
        except Exception as e:
            self._log(f"错误: {e}")
        finally:
            QTimer.singleShot(0, self._done)

    def _done(self):
        self.run_btn.setEnabled(True)
        self.run_btn.setText("开始获取")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, handlers=[logging.StreamHandler()])
    app = QApplication(sys.argv)
    w = UserListFetcher()
    w.show()
    sys.exit(app.exec())
